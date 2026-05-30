"""
Classe base abstrata para geradores de formulário de contribuição.

Todos os geradores específicos (artigo, livro, etc.) devem herdar
desta classe e implementar os métodos abstratos.
"""

import io
import logging
import html
from abc import ABC, abstractmethod
from datetime import datetime

logger = logging.getLogger(__name__)


class FieldDefinition:
    """
    Define um campo do formulário de contribuição.

    Args:
        label: Nome/rótulo do campo
        field_type: Tipo ('text', 'textarea', 'date', 'number', 'url',
                          'choice', 'boolean', 'email')
        required: Se o campo é obrigatório
        instruction: Instrução detalhada de preenchimento
        options: Lista de opções válidas (para field_type='choice')
        example: Exemplo de valor preenchido
    """

    def __init__(self, label, field_type='text', required=True,
                 instruction='', options=None, example=''):
        self.label = label
        self.field_type = field_type
        self.required = required
        self.instruction = instruction
        self.options = options or []
        self.example = example

    @property
    def required_label(self):
        return '* Obrigatório' if self.required else 'Opcional'


class BaseFormGenerator(ABC):
    """
    Gerador base para formulários de contribuição de conteúdo.

    Subclasses devem implementar:
        - get_form_title()
        - get_form_description()
        - get_fields()
        - get_instructions()
    """

    # Paleta de cores para o XLSX — override nas subclasses para personalizar
    COLOR_HEADER_BG = "1A3A5C"    # Azul escuro
    COLOR_HEADER_FG = "FFFFFF"    # Branco
    COLOR_REQUIRED_BG = "FFF3CD"  # Amarelo claro (campos obrigatórios)
    COLOR_OPTIONAL_BG = "F8F9FA"  # Cinza claro (campos opcionais)
    COLOR_SECTION_BG = "E8F4F8"   # Azul suave (seções)
    COLOR_INSTRUCTION_BG = "EAF7EA"  # Verde suave (instruções)
    COLOR_ACCENT = "2E86AB"       # Azul médio

    @abstractmethod
    def get_form_title(self) -> str:
        """Retorna o título do formulário."""
        pass

    @abstractmethod
    def get_form_description(self) -> str:
        """Retorna a descrição/objetivo do formulário."""
        pass

    @abstractmethod
    def get_fields(self) -> list:
        """Retorna lista de FieldDefinition para os campos do formulário."""
        pass

    @abstractmethod
    def get_instructions(self) -> list:
        """Retorna lista de tuplas (título, texto) com instruções de uso."""
        pass

    def get_filename_base(self) -> str:
        """Nome base para o arquivo de download (sem extensão)."""
        today = datetime.now().strftime('%Y%m%d')
        safe_title = self.get_form_title().lower().replace(' ', '_')
        return f"formulario_contribuicao_{safe_title}_{today}"

    # ─────────────────────────────────────────────
    # Geração XLSX
    # ─────────────────────────────────────────────

    def generate_xlsx(self) -> io.BytesIO:
        """Gera formulário em formato XLSX com formatação rica."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
            except ImportError:
                from openpyxl.data_validation import DataValidation
        except ImportError:
            logger.error("openpyxl não instalado. Execute: pip install openpyxl")
            raise

        wb = Workbook()

        # ── Aba 1: Formulário ──────────────────────────────────────────────
        ws_form = wb.active
        ws_form.title = "📝 Formulário"

        # Estilos reutilizáveis
        def make_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def make_border(style='thin'):
            side = Side(style=style)
            return Border(left=side, right=side, top=side, bottom=side)

        header_font = Font(name='Calibri', bold=True, color=self.COLOR_HEADER_FG, size=11)
        title_font = Font(name='Calibri', bold=True, size=14, color=self.COLOR_ACCENT)
        label_font = Font(name='Calibri', bold=True, size=10)
        normal_font = Font(name='Calibri', size=10)
        small_font = Font(name='Calibri', size=9, italic=True, color="888888")
        required_font = Font(name='Calibri', bold=True, size=9, color="CC0000")

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

        header_fill = make_fill(self.COLOR_HEADER_BG)
        required_fill = make_fill(self.COLOR_REQUIRED_BG)
        optional_fill = make_fill(self.COLOR_OPTIONAL_BG)
        section_fill = make_fill(self.COLOR_SECTION_BG)
        border = make_border()

        # Configurar colunas
        ws_form.column_dimensions['A'].width = 28  # Campo
        ws_form.column_dimensions['B'].width = 50  # Seu preenchimento
        ws_form.column_dimensions['C'].width = 15  # Obrigatório?
        ws_form.column_dimensions['D'].width = 55  # Instruções

        # Linha 1: Título principal
        ws_form.merge_cells('A1:D1')
        title_cell = ws_form['A1']
        title_cell.value = f"🗒️ {self.get_form_title()}"
        title_cell.font = title_font
        title_cell.fill = make_fill(self.COLOR_SECTION_BG)
        title_cell.alignment = center_align
        title_cell.border = border
        ws_form.row_dimensions[1].height = 36

        # Linha 2: Descrição
        ws_form.merge_cells('A2:D2')
        desc_cell = ws_form['A2']
        desc_cell.value = self.get_form_description()
        desc_cell.font = small_font
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_form.row_dimensions[2].height = 40

        # Linha 3: Cabeçalhos das colunas
        headers = ['Campo', 'Seu Preenchimento', 'Obrigatório?', 'Instruções de Preenchimento']
        for col, header in enumerate(headers, 1):
            cell = ws_form.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        ws_form.row_dimensions[3].height = 22

        # Linha 4: Aviso
        ws_form.merge_cells('A4:D4')
        aviso_cell = ws_form['A4']
        aviso_cell.value = (
            "⚠️  Preencha APENAS a coluna 'Seu Preenchimento' (coluna B). "
            "Campos marcados com * são obrigatórios."
        )
        aviso_cell.font = Font(name='Calibri', bold=True, size=9, color="856404")
        aviso_cell.fill = make_fill("FFF3CD")
        aviso_cell.alignment = left_align
        ws_form.row_dimensions[4].height = 20

        # Dados dos campos
        row = 5
        fields = self.get_fields()
        for field in fields:
            fill = required_fill if field.required else optional_fill

            # Coluna A: Label
            cell_a = ws_form.cell(row=row, column=1, value=field.label)
            cell_a.font = label_font
            cell_a.fill = fill
            cell_a.alignment = left_align
            cell_a.border = border

            # Coluna B: Preenchimento pelo usuário
            example_hint = f"[Exemplo: {field.example}]" if field.example else ""
            cell_b = ws_form.cell(row=row, column=2, value=example_hint)
            cell_b.font = Font(name='Calibri', size=10, italic=True, color="AAAAAA")
            cell_b.fill = make_fill("FFFFFF")
            cell_b.alignment = left_align
            cell_b.border = border

            # Validação de lista suspensa para campos 'choice'
            if field.field_type == 'choice' and field.options:
                options_str = '","'.join(field.options)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{options_str}"',
                    allow_blank=not field.required,
                    showDropDown=False,
                    showErrorMessage=True,
                    error=f"Selecione uma das opções válidas para '{field.label}'.",
                    errorTitle="Valor inválido"
                )
                ws_form.add_data_validation(dv)
                # Usar string de coordenada (compatível com openpyxl >= 3.1)
                cell_coord = ws_form.cell(row=row, column=2).coordinate
                dv.sqref = cell_coord


            # Coluna C: Obrigatório?
            cell_c = ws_form.cell(row=row, column=3, value=field.required_label)
            cell_c.font = required_font if field.required else small_font
            cell_c.fill = fill
            cell_c.alignment = center_align
            cell_c.border = border

            # Coluna D: Instrução
            cell_d = ws_form.cell(row=row, column=4, value=field.instruction)
            cell_d.font = small_font
            cell_d.fill = make_fill(self.COLOR_INSTRUCTION_BG)
            cell_d.alignment = left_align
            cell_d.border = border

            ws_form.row_dimensions[row].height = 40
            row += 1

        # Linha final: Assinatura/envio
        ws_form.merge_cells(f'A{row}:D{row}')
        end_cell = ws_form[f'A{row}']
        end_cell.value = (
            "✅ Após preencher, salve o arquivo e envie para a equipe editorial da plataforma. "
            "Formulário gerado automaticamente pela plataforma CG Bookstore."
        )
        end_cell.font = Font(name='Calibri', size=9, italic=True, color="666666")
        end_cell.fill = make_fill(self.COLOR_SECTION_BG)
        end_cell.alignment = left_align
        ws_form.row_dimensions[row].height = 25

        # ── Aba 2: Instruções ─────────────────────────────────────────────
        ws_inst = wb.create_sheet(title="📋 Instruções")
        ws_inst.column_dimensions['A'].width = 30
        ws_inst.column_dimensions['B'].width = 80

        ws_inst.merge_cells('A1:B1')
        inst_title = ws_inst['A1']
        inst_title.value = f"📋 Instruções — {self.get_form_title()}"
        inst_title.font = title_font
        inst_title.fill = make_fill(self.COLOR_SECTION_BG)
        inst_title.alignment = center_align
        ws_inst.row_dimensions[1].height = 30

        row_i = 2
        for titulo, texto in self.get_instructions():
            # Título da instrução
            ws_inst.merge_cells(f'A{row_i}:B{row_i}')
            cell_t = ws_inst[f'A{row_i}']
            cell_t.value = titulo
            cell_t.font = Font(name='Calibri', bold=True, size=11, color=self.COLOR_ACCENT)
            cell_t.fill = make_fill(self.COLOR_SECTION_BG)
            cell_t.alignment = left_align
            ws_inst.row_dimensions[row_i].height = 22
            row_i += 1

            # Texto da instrução
            ws_inst.merge_cells(f'A{row_i}:B{row_i}')
            cell_txt = ws_inst[f'A{row_i}']
            cell_txt.value = texto
            cell_txt.font = Font(name='Calibri', size=10)
            cell_txt.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws_inst.row_dimensions[row_i].height = 60
            row_i += 1
            row_i += 1  # Espaço

        # Proteger abas (opcional: impede edição fora da coluna B)
        # ws_form.protection.sheet = True  # Descomentado futuramente

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ─────────────────────────────────────────────
    # Geração TXT
    # ─────────────────────────────────────────────

    def generate_txt(self) -> io.BytesIO:
        """Gera formulário em texto simples (fallback sem dependências extras)."""
        lines = []
        separator = "=" * 70
        thin_sep = "-" * 70

        lines.append(separator)
        lines.append(f"  {self.get_form_title().upper()}")
        lines.append(separator)
        lines.append(self.get_form_description())
        lines.append("")
        lines.append("ATENÇÃO: Preencha apenas os campos após o sinal '>>>'.")
        lines.append("Campos marcados com [*] são obrigatórios.")
        lines.append(separator)
        lines.append("")

        for field in self.get_fields():
            required_mark = "[*] " if field.required else "[ ] "
            lines.append(f"{required_mark}{field.label.upper()}")
            if field.instruction:
                lines.append(f"  Instrução: {field.instruction}")
            if field.example:
                lines.append(f"  Exemplo: {field.example}")
            if field.options:
                lines.append(f"  Opções válidas: {', '.join(field.options)}")
            lines.append(f"  >>> ")
            lines.append(thin_sep)
            lines.append("")

        lines.append(separator)
        for titulo, texto in self.get_instructions():
            lines.append(f"### {titulo}")
            lines.append(texto)
            lines.append("")
        lines.append(separator)
        lines.append("Formulário gerado automaticamente pela plataforma CG Bookstore.")
        lines.append(separator)

        content = "\n".join(lines)
        buffer = io.BytesIO(content.encode('utf-8'))
        buffer.seek(0)
        return buffer

    # ─────────────────────────────────────────────
    # Geração PDF
    # ─────────────────────────────────────────────

    def generate_pdf(self) -> io.BytesIO:
        """Gera formulário em PDF com formatação profissional via reportlab."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table,
                TableStyle, HRFlowable, PageBreak
            )
        except ImportError:
            logger.error("reportlab não instalado. Execute: pip install reportlab")
            raise

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
            title=self.get_form_title(),
            author="CG Bookstore",
        )

        styles = getSampleStyleSheet()

        # Estilos personalizados
        COLOR_PRIMARY = colors.HexColor("#1A3A5C")
        COLOR_ACCENT = colors.HexColor("#2E86AB")
        COLOR_REQUIRED_BG = colors.HexColor("#FFF3CD")
        COLOR_OPTIONAL_BG = colors.HexColor("#F8F9FA")
        COLOR_HEADER_BG = colors.HexColor("#1A3A5C")
        COLOR_INSTR_BG = colors.HexColor("#EAF7EA")

        style_title = ParagraphStyle(
            'TitleCustom', parent=styles['Title'],
            fontName='Helvetica-Bold', fontSize=18,
            textColor=COLOR_PRIMARY, spaceAfter=6
        )
        style_subtitle = ParagraphStyle(
            'SubtitleCustom', parent=styles['Normal'],
            fontName='Helvetica', fontSize=10,
            textColor=colors.HexColor("#555555"), spaceAfter=14
        )
        style_label = ParagraphStyle(
            'LabelCustom', parent=styles['Normal'],
            fontName='Helvetica-Bold', fontSize=9,
            textColor=COLOR_PRIMARY
        )
        style_instruction = ParagraphStyle(
            'InstrCustom', parent=styles['Normal'],
            fontName='Helvetica', fontSize=8,
            textColor=colors.HexColor("#555555")
        )
        style_section = ParagraphStyle(
            'SectionCustom', parent=styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=12,
            textColor=COLOR_ACCENT, spaceBefore=14, spaceAfter=6
        )
        style_body = ParagraphStyle(
            'BodyCustom', parent=styles['Normal'],
            fontName='Helvetica', fontSize=9,
            textColor=colors.HexColor("#333333"), spaceAfter=4
        )

        story = []

        # Cabeçalho
        story.append(Paragraph(f"🗒️ {self.get_form_title()}", style_title))
        story.append(Paragraph(self.get_form_description(), style_subtitle))
        story.append(HRFlowable(width="100%", thickness=2, color=COLOR_ACCENT))
        story.append(Spacer(1, 0.4 * cm))

        # Aviso
        aviso_data = [[
            Paragraph(
                "⚠️  Preencha os campos abaixo com as informações do conteúdo que deseja "
                "contribuir. Campos marcados com * são obrigatórios.",
                style_instruction
            )
        ]]
        aviso_table = Table(aviso_data, colWidths=[16.5 * cm])
        aviso_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFF3CD")),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#856404")),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(aviso_table)
        story.append(Spacer(1, 0.5 * cm))

        # Tabela de campos
        header_row = [
            Paragraph('<font color="white"><b>Campo</b></font>', style_label),
            Paragraph('<font color="white"><b>Preencha Aqui</b></font>', style_label),
            Paragraph('<font color="white"><b>Obrig.?</b></font>', style_label),
            Paragraph('<font color="white"><b>Instruções</b></font>', style_label),
        ]

        table_data = [header_row]
        row_colors = []

        for i, field in enumerate(self.get_fields()):
            bg = COLOR_REQUIRED_BG if field.required else COLOR_OPTIONAL_BG
            row_colors.append(bg)

            req_text = "✓ Obrig." if field.required else "Opcional"
            req_color = "#CC0000" if field.required else "#666666"
            example_note = f"<br/><i><font color='#AAAAAA'>[Ex: {field.example}]</font></i>" if field.example else ""

            row = [
                Paragraph(f'<b>{field.label}</b>', style_label),
                Paragraph(example_note, style_instruction),
                Paragraph(f'<font color="{req_color}"><b>{req_text}</b></font>', style_instruction),
                Paragraph(html.escape(field.instruction or "—"), style_instruction),
            ]
            table_data.append(row)

        col_widths = [4 * cm, 4.5 * cm, 2 * cm, 6 * cm]
        fields_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        table_style = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_HEADER_BG),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            # Bordas
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ('BOX', (0, 0), (-1, -1), 1, COLOR_PRIMARY),
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]

        # Cores alternadas das linhas
        for i, bg in enumerate(row_colors):
            table_style.append(('BACKGROUND', (0, i + 1), (-1, i + 1), bg))

        fields_table.setStyle(TableStyle(table_style))
        story.append(fields_table)
        story.append(Spacer(1, 0.8 * cm))

        # ── Seção de instruções ───────────────────────────────────────────
        story.append(PageBreak())
        story.append(Paragraph("📋 Instruções e Políticas da Plataforma", style_section))
        story.append(HRFlowable(width="100%", thickness=1, color=COLOR_ACCENT))
        story.append(Spacer(1, 0.3 * cm))

        for titulo, texto in self.get_instructions():
            story.append(Paragraph(titulo, style_section))
            # Converter \n para <br/> e escapar HTML para evitar erros de parsing no reportlab
            texto_safe = html.escape(texto).replace('\n', '<br/>')
            instr_data = [[Paragraph(texto_safe, style_body)]]
            instr_table = Table(instr_data, colWidths=[16.5 * cm])
            instr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), COLOR_INSTR_BG),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ]))
            story.append(instr_table)
            story.append(Spacer(1, 0.4 * cm))

        # Rodapé
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CCCCCC")))
        story.append(Paragraph(
            f"Formulário gerado automaticamente em "
            f"{datetime.now().strftime('%d/%m/%Y às %H:%M')} pela plataforma CG Bookstore.",
            style_instruction
        ))

        doc.build(story)
        buffer.seek(0)
        return buffer
